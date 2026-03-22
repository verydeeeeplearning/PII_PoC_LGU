"""PoC 결과 Excel 리포트 작성기 (6-sheet).

openpyxl 직접 사용 - pandas/openpyxl/numpy 외 외부 패키지 없음.

Sheet 구성:
    1. 요약           - 데이터 조건, split 건수, 핵심 지표, PoC 판정
    2. 데이터 통계    - TP/FP 비율, 월별 분포, fp_description 카테고리, dedup
    3. 모델 성능      - 3종 Split 비교표, PASS/FAIL 색상
    4. Coverage 곡선  - τ 테이블, 권장 τ 하이라이트
    5. Rule 기여도    - rule_id별 히트율/정밀도
    6. 오분류 분석    - 패턴 상위 15 + 샘플 200건
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False


# ── 색상 상수 (AARRGGBB 8자리 - openpyxl 표준) ────────────────────────────────
COLOR_PASS      = "FF00B050"   # 녹색
COLOR_FAIL      = "FFFF0000"   # 빨강
COLOR_HIGHLIGHT = "FFFFFF00"   # 노랑 (권장 τ)
COLOR_HEADER    = "FF4472C4"   # 파랑 (헤더)
COLOR_SUBHEADER = "FFD9E1F2"   # 연파랑 (소제목)
COLOR_WHITE     = "FFFFFFFF"


# ── 데이터 컨테이너 ────────────────────────────────────────────────────────────

@dataclass
class PocReportData:
    """PocExcelWriter에 전달할 데이터 컨테이너."""
    # Sheet 1 - 요약
    data_condition: str = "Label Only"          # "Label Only" / "Label + Sumologic"
    split_summary: dict = field(default_factory=dict)   # {train_n, test_n, split_method}
    poc_criteria: dict = field(default_factory=dict)    # check_poc_criteria() 반환

    # Sheet 2 - 데이터 통계
    binary_stats: dict = field(default_factory=dict)            # compute_binary_stats() 반환
    class_imbalance: pd.DataFrame = field(default_factory=pd.DataFrame)
    dedup_before: int = 0
    dedup_after: int = 0
    fp_description_stats: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 3 - 모델 성능
    split_comparison: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 4 - Coverage 곡선
    coverage_curve: dict = field(default_factory=dict)   # compute_coverage_precision_curve() 반환

    # Sheet 5 - Rule 기여도
    rule_contribution: pd.DataFrame = field(default_factory=pd.DataFrame)
    class_rule_contribution: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 6 - 오분류 분석
    error_patterns: list = field(default_factory=list)   # [(actual, predicted, count)]
    error_samples: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 1 보강 - 실행 메타데이터 + 비즈니스 임팩트
    run_metadata: dict = field(default_factory=dict)
    business_impact: dict = field(default_factory=dict)

    # Sheet 2 보강 - 조직별 분포
    org_stats: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 3 보강 - 클래스별 성능
    class_metrics: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 5 보강 - Rule vs ML Coverage
    rule_vs_ml_coverage: dict = field(default_factory=dict)

    # Sheet 6 보강 - 오분류 위험도
    error_risk_summary: dict = field(default_factory=dict)

    # Sheet 7 - ML 예측 신뢰도 분포
    confidence_distribution: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 8 - Feature Importance (run_report.py 통합)
    feature_importance_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    feature_group_importance: dict = field(default_factory=dict)

    # Sheet 9 - 데이터 진단 (run_report.py --include-diagnosis)
    column_risk_registry: pd.DataFrame = field(default_factory=pd.DataFrame)
    split_robustness: pd.DataFrame = field(default_factory=pd.DataFrame)
    ablation_results: pd.DataFrame = field(default_factory=pd.DataFrame)

    # Sheet 10 - Decision Combiner 평가
    dc_eval_result: dict = field(default_factory=dict)


# ── 메인 작성기 ────────────────────────────────────────────────────────────────

class PocExcelWriter:
    """6-sheet PoC 결과 Excel 파일 작성기.

    Usage::

        data = PocReportData(...)
        writer = PocExcelWriter(data)
        writer.write(Path("outputs/poc_report.xlsx"))
    """

    def __init__(self, data: PocReportData) -> None:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl이 설치되어 있지 않습니다.")
        self.data = data

    def write(self, output_path: Path) -> None:
        """9-sheet Excel 파일을 생성한다."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # 기본 시트 제거
        if wb.active:
            wb.remove(wb.active)

        self._build_summary_sheet(wb.create_sheet("1_요약"))
        self._build_data_stats_sheet(wb.create_sheet("2_데이터통계"))
        self._build_model_performance_sheet(wb.create_sheet("3_모델성능"))
        self._build_coverage_curve_sheet(wb.create_sheet("4_Coverage곡선"))
        self._build_rule_contribution_sheet(wb.create_sheet("5_Rule기여도"))
        self._build_error_analysis_sheet(wb.create_sheet("6_오분류분석"))
        self._build_confidence_sheet(wb.create_sheet("7_신뢰도분포"))
        self._build_feature_importance_sheet(wb.create_sheet("8_FeatureImportance"))
        self._build_diagnosis_sheet(wb.create_sheet("9_데이터진단"))
        self._build_decision_combiner_sheet(wb.create_sheet("10_DecisionCombiner"))

        wb.save(str(output_path))

    # ── Sheet 1: 요약 ──────────────────────────────────────────────────────────

    def _build_summary_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "PoC 결과 요약")

        poc = self.data.poc_criteria
        split_sum = self.data.split_summary

        # 데이터 조건
        kv_rows = [
            ("데이터 조건", self.data.data_condition),
            ("학습 데이터 건수", split_sum.get("train_n", "-")),
            ("테스트 데이터 건수", split_sum.get("test_n", "-")),
            ("Split 방법", split_sum.get("split_method", "-")),
        ]
        row = self._write_kv_rows(ws, kv_rows, start_row=3)

        # PoC 판정 섹션
        row += 2
        self._write_section_title(ws, row, 1, "PoC 판정 기준")
        row += 1

        # 숫자형이면 포맷, 아니면 그대로
        def _fmt(v):
            try:
                return f"{float(v):.4f}"
            except (TypeError, ValueError):
                return str(v) if v is not None else "-"

        criteria_rows = [
            ("F1-macro",           _fmt(poc.get("f1_macro",     "-")), "≥ 0.70"),
            ("TP Recall",          _fmt(poc.get("tp_recall",    "-")), "≥ 0.75"),
            ("FP Precision",       _fmt(poc.get("fp_precision", "-")), "≥ 0.85"),
        ]

        # 헤더
        for col_offset, hdr in enumerate(["지표", "값", "기준"], start=1):
            cell = ws.cell(row=row, column=col_offset, value=hdr)
            self._color_cell(cell, COLOR_HEADER)
            cell.font = Font(color="FFFFFFFF", bold=True)
        row += 1

        for metric, value, threshold in criteria_rows:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=threshold)
            row += 1

        # 종합 판정
        row += 1
        passes = poc.get("passes", False)
        verdict = "PASS" if passes else "FAIL"
        verdict_cell = ws.cell(row=row, column=1, value=f"종합 판정: {verdict}")
        verdict_cell.font = Font(bold=True, size=14)
        self._color_cell(verdict_cell, COLOR_PASS if passes else COLOR_FAIL)
        if not passes:
            verdict_cell.font = Font(bold=True, size=14, color="FFFFFFFF")
        row += 1

        # 실행 메타데이터 섹션
        meta = self.data.run_metadata
        if meta:
            row += 1
            self._write_section_title(ws, row, 1, "실행 메타데이터")
            row += 1
            meta_rows = [
                ("실행 일시", meta.get("run_datetime", "-")),
                ("모델 경로", meta.get("model_path", "-")),
                ("데이터 기간", meta.get("data_date_range", "-")),
            ]
            row = self._write_kv_rows(ws, meta_rows, start_row=row)

        # 비즈니스 임팩트 섹션
        bi = self.data.business_impact
        if bi:
            row += 1
            self._write_section_title(ws, row, 1, "비즈니스 임팩트 (권장 τ 적용 시)")
            row += 1
            goal_met = bi.get("phase1_goal_40pct_met", False)
            bi_rows = [
                ("전체 FP 건수", bi.get("total_fp", "-")),
                ("자동 처리 Coverage", f"{bi.get('coverage_at_target', 0):.1%}"),
                ("자동 처리 추정 건수", bi.get("estimated_auto_fp", "-")),
                ("Phase 1 목표 달성 (≥40%)", "PASS" if goal_met else "FAIL"),
            ]
            row = self._write_kv_rows(ws, bi_rows, start_row=row)
            goal_cell = ws.cell(row=row - 1, column=2)
            self._color_cell(goal_cell, COLOR_PASS if goal_met else COLOR_FAIL)
            if not goal_met:
                goal_cell.font = Font(color=COLOR_WHITE)

        self._auto_col_width(ws)

    # ── Sheet 2: 데이터 통계 ───────────────────────────────────────────────────

    def _build_data_stats_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "데이터 통계")
        row = 3

        # TP/FP 집계
        total = self.data.binary_stats.get("total", {})
        kv_rows = [
            ("전체 건수",   total.get("total",    0)),
            ("TP 건수",     total.get("tp",       0)),
            ("FP 건수",     total.get("fp",       0)),
            ("TP 비율",     f"{total.get('tp_ratio', 0.0):.2%}"),
            ("FP 비율",     f"{total.get('fp_ratio', 0.0):.2%}"),
            ("중복 제거 전", self.data.dedup_before),
            ("중복 제거 후", self.data.dedup_after),
        ]
        row = self._write_kv_rows(ws, kv_rows, start_row=row)

        # 월별 분포
        by_month = self.data.binary_stats.get("by_month", pd.DataFrame())
        if not isinstance(by_month, pd.DataFrame):
            by_month = pd.DataFrame()
        if not by_month.empty:
            row += 2
            self._write_section_title(ws, row, 1, "월별 TP/FP 분포")
            row += 1
            row = self._write_table(ws, by_month, row, 1, COLOR_SUBHEADER)

        # 클래스 불균형
        if not self.data.class_imbalance.empty:
            row += 2
            self._write_section_title(ws, row, 1, "클래스별 분포")
            row += 1
            row = self._write_table(ws, self.data.class_imbalance, row, 1, COLOR_SUBHEADER)

        # fp_description 카테고리
        if not self.data.fp_description_stats.empty:
            row += 2
            self._write_section_title(ws, row, 1, "FP Description 카테고리")
            row += 1
            row = self._write_table(ws, self.data.fp_description_stats, row, 1, COLOR_SUBHEADER)

        # 조직별 분포
        if not self.data.org_stats.empty:
            row += 2
            self._write_section_title(ws, row, 1, "조직별 TP/FP 분포")
            row += 1
            row = self._write_table(ws, self.data.org_stats, row, 1, COLOR_SUBHEADER)

        self._auto_col_width(ws)

    # ── Sheet 3: 모델 성능 ─────────────────────────────────────────────────────

    def _build_model_performance_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "모델 성능 - 3종 Split 비교")
        row = 3

        if self.data.split_comparison.empty:
            ws.cell(row=row, column=1, value="(데이터 없음)")
            return

        df = self.data.split_comparison
        # 헤더
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=str(col_name))
            self._color_cell(cell, COLOR_HEADER)
            cell.font = Font(color="FFFFFFFF", bold=True)
        row += 1

        for _, data_row in df.iterrows():
            for col_idx, (col_name, val) in enumerate(data_row.items(), start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                # PASS/FAIL 색상
                if col_name == "poc_verdict":
                    if val == "PASS":
                        self._color_cell(cell, COLOR_PASS)
                    elif val == "FAIL":
                        self._color_cell(cell, COLOR_FAIL)
                        cell.font = Font(color=COLOR_WHITE)
            row += 1

        # 클래스별 성능 섹션
        if not self.data.class_metrics.empty:
            row += 2
            self._write_section_title(ws, row, 1, "클래스별 Precision / Recall / F1")
            row += 1
            row = self._write_table(ws, self.data.class_metrics, row, 1, COLOR_SUBHEADER)

        self._auto_col_width(ws)

    # ── Sheet 4: Coverage 곡선 ─────────────────────────────────────────────────

    def _build_coverage_curve_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "Coverage-Precision 곡선 (τ 스윕)")
        row = 3

        curve = self.data.coverage_curve
        if not isinstance(curve, dict):
            ws.cell(row=row, column=1, value="(데이터 없음)")
            return

        curve_df = curve.get("curve", pd.DataFrame())
        recommended_tau = curve.get("recommended_tau", None)

        if not isinstance(curve_df, pd.DataFrame) or curve_df.empty:
            ws.cell(row=row, column=1, value="(커브 데이터 없음)")
            return

        # 권장 τ 안내
        if recommended_tau is not None:
            ws.cell(row=row, column=1, value=f"권장 τ: {recommended_tau:.4f}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        # 테이블
        for col_idx, col_name in enumerate(curve_df.columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=str(col_name))
            self._color_cell(cell, COLOR_HEADER)
            cell.font = Font(color="FFFFFFFF", bold=True)
        row += 1

        for _, data_row in curve_df.iterrows():
            tau_val = data_row.get("tau", None)
            for col_idx, val in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
            # 권장 τ 행 하이라이트
            if recommended_tau is not None and tau_val == recommended_tau:
                for col_idx in range(1, len(curve_df.columns) + 1):
                    self._color_cell(ws.cell(row=row, column=col_idx), COLOR_HIGHLIGHT)
            row += 1

        self._auto_col_width(ws)

    # ── Sheet 5: Rule 기여도 ───────────────────────────────────────────────────

    def _build_rule_contribution_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "Rule 기여도 분석")
        row = 3

        if not self.data.rule_contribution.empty:
            self._write_section_title(ws, row, 1, "룰별 히트율 / 정밀도")
            row += 1
            row = self._write_table(ws, self.data.rule_contribution, row, 1, COLOR_HEADER)

        if not self.data.class_rule_contribution.empty:
            row += 2
            self._write_section_title(ws, row, 1, "클래스별 Rule 기여도")
            row += 1
            row = self._write_table(ws, self.data.class_rule_contribution, row, 1, COLOR_SUBHEADER)

        # Rule vs ML Coverage 비교
        rvc = self.data.rule_vs_ml_coverage
        if rvc:
            row += 2
            self._write_section_title(ws, row, 1, "Rule vs ML Coverage 비교")
            row += 1
            rvc_rows = [
                ("Rule 단독 Coverage",  f"{rvc.get('rule_only_coverage', 0):.1%}"),
                ("ML 총 Coverage",      f"{rvc.get('ml_total_coverage', 0):.1%}"),
                ("ML 추가 Coverage",    f"{rvc.get('ml_additional_coverage', 0):.1%}"),
                ("중복 검출 비율",      f"{rvc.get('overlap_rate', 0):.1%}"),
            ]
            row = self._write_kv_rows(ws, rvc_rows, start_row=row)

        self._auto_col_width(ws)

    # ── Sheet 6: 오분류 분석 ───────────────────────────────────────────────────

    def _build_error_analysis_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "오분류 분석")
        row = 3

        # 오분류 위험도 요약 (먼저 출력)
        ers = self.data.error_risk_summary
        if ers:
            self._write_section_title(ws, row, 1, "오분류 위험도 요약")
            row += 1
            ers_rows = [
                ("전체 오분류 건수",          ers.get("total_errors", 0)),
                ("위험 오류 (FP->TP 통과)",    ers.get("fp_to_tp_count", 0)),
                ("누락 오류 (TP->FP 처리)",    ers.get("tp_to_fp_count", 0)),
                ("위험 오류 비율",             f"{ers.get('risk_rate', 0):.1%}"),
            ]
            row = self._write_kv_rows(ws, ers_rows, start_row=row)
            # fp_to_tp_count 셀 색상: 0이면 녹색, >0이면 빨강
            fp_to_tp_val = ers.get("fp_to_tp_count", 0)
            fp_to_tp_cell = ws.cell(row=row - 3, column=2)
            self._color_cell(fp_to_tp_cell, COLOR_PASS if fp_to_tp_val == 0 else COLOR_FAIL)
            if fp_to_tp_val > 0:
                fp_to_tp_cell.font = Font(color=COLOR_WHITE)
            row += 2

        # 상위 15 패턴
        if self.data.error_patterns:
            self._write_section_title(ws, row, 1, "주요 오분류 패턴 (상위 15)")
            row += 1
            headers = ["실제 클래스", "예측 클래스", "건수"]
            for col_idx, hdr in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=hdr)
                self._color_cell(cell, COLOR_HEADER)
                cell.font = Font(color="FFFFFFFF", bold=True)
            row += 1
            for actual, predicted, count in self.data.error_patterns[:15]:
                ws.cell(row=row, column=1, value=str(actual))
                ws.cell(row=row, column=2, value=str(predicted))
                ws.cell(row=row, column=3, value=int(count))
                row += 1

        # 샘플 200건
        if not self.data.error_samples.empty:
            row += 2
            self._write_section_title(ws, row, 1, "오분류 샘플 (상위 200건)")
            row += 1
            sample_df = self.data.error_samples.head(200)
            row = self._write_table(ws, sample_df, row, 1, COLOR_SUBHEADER)

        self._auto_col_width(ws)

    # ── Sheet 7: ML 신뢰도 분포 ───────────────────────────────────────────────

    def _build_confidence_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "ML 예측 신뢰도 분포 (τ 구간별)")
        row = 3

        if self.data.confidence_distribution.empty:
            ws.cell(row=row, column=1, value="(ML 모델 없음 또는 --skip-ml)")
            return

        row = self._write_table(ws, self.data.confidence_distribution, row, 1, COLOR_HEADER)
        self._auto_col_width(ws)

    # ── Sheet 8: Feature Importance ──────────────────────────────────────────

    def _build_feature_importance_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "Feature Importance")

        fi = self.data.feature_importance_df
        groups = self.data.feature_group_importance

        if fi.empty:
            ws.cell(row=3, column=1, value="(Feature Importance 데이터 없음)")
            return

        # 그룹별 합산 테이블
        row = 3
        self._write_section_title(ws, row, 1, "그룹별 Importance 합산")
        row += 1

        if groups:
            total = sum(groups.values()) or 1
            group_df = pd.DataFrame([
                {"그룹": g, "Importance": v, "비율": f"{v/total:.1%}"}
                for g, v in sorted(groups.items(), key=lambda x: -x[1])
            ])
            row = self._write_table(ws, group_df, row, 1, COLOR_HEADER)

        # 상위 30 피처 테이블
        row += 2
        self._write_section_title(ws, row, 1, "상위 30 Feature")
        row += 1
        top_n = fi.head(30).copy()
        top_n.insert(0, "Rank", range(1, len(top_n) + 1))
        row = self._write_table(ws, top_n, row, 1, COLOR_HEADER)

        self._auto_col_width(ws)

    # ── Sheet 9: 데이터 진단 ──────────────────────────────────────────────────

    def _build_diagnosis_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "데이터 진단")

        registry = self.data.column_risk_registry
        robustness = self.data.split_robustness
        ablation = self.data.ablation_results

        has_data = (
            not registry.empty
            or not robustness.empty
            or not ablation.empty
        )

        if not has_data:
            ws.cell(row=3, column=1, value="(진단 미실행 - --include-diagnosis로 활성화)")
            return

        row = 3

        # Column Risk Registry
        if not registry.empty:
            self._write_section_title(ws, row, 1, "Column Risk Registry")
            row += 1
            row = self._write_table(ws, registry, row, 1, COLOR_HEADER)
            row += 2

        # Split Robustness
        if not robustness.empty:
            self._write_section_title(ws, row, 1, "Split Robustness (split별 F1 비교)")
            row += 1
            row = self._write_table(ws, robustness, row, 1, COLOR_HEADER)
            row += 2

        # Feature Ablation
        if not ablation.empty:
            self._write_section_title(ws, row, 1, "Feature Block Ablation")
            row += 1
            row = self._write_table(ws, ablation, row, 1, COLOR_HEADER)

        self._auto_col_width(ws)

    # ── Sheet 10: Decision Combiner ──────────────────────────────────────────

    def _build_decision_combiner_sheet(self, ws) -> None:
        self._write_section_title(ws, 1, 1, "Decision Combiner 평가 (RULE + ML 결합)")

        dc = self.data.dc_eval_result
        if not dc:
            ws.cell(row=3, column=1, value="(Decision Combiner 시뮬레이션 미실행)")
            return

        row = 3

        # F1 비교
        self._write_section_title(ws, row, 1, "F1-macro 비교")
        row += 1
        kv_rows = [
            ("ML 단독 F1-macro", f"{dc.get('ml_f1', 0):.4f}"),
            ("Decision Combiner F1-macro", f"{dc.get('dc_f1', 0):.4f}"),
            ("차이 (DC − ML)", f"{dc.get('dc_f1', 0) - dc.get('ml_f1', 0):+.4f}"),
            ("평가 샘플 수", f"{dc.get('total_samples', 0):,}"),
        ]
        row = self._write_kv_rows(ws, kv_rows, start_row=row)
        row += 2

        # 4분면 Confusion Matrix
        cm_df = dc.get("confusion_matrix")
        if cm_df is not None and not cm_df.empty:
            self._write_section_title(ws, row, 1, "4분면 (Confusion Matrix)")
            row += 1
            row = self._write_table(ws, cm_df.reset_index().rename(
                columns={"index": ""}
            ), row, 1, COLOR_HEADER)
            row += 2

        # Decision Source 분포
        src_dist = dc.get("decision_source_dist", {})
        if src_dist:
            self._write_section_title(ws, row, 1, "Decision Source 분포")
            row += 1
            src_df = pd.DataFrame([
                {"Source": k, "건수": v, "비율": f"{v / dc.get('total_samples', 1):.1%}"}
                for k, v in sorted(src_dist.items(), key=lambda x: -x[1])
            ])
            row = self._write_table(ws, src_df, row, 1, COLOR_HEADER)
            row += 2

        # Reason Code 분포
        reason_dist = dc.get("reason_code_dist", {})
        if reason_dist:
            self._write_section_title(ws, row, 1, "Reason Code 분포")
            row += 1
            reason_df = pd.DataFrame([
                {"Reason": k, "건수": v, "비율": f"{v / dc.get('total_samples', 1):.1%}"}
                for k, v in sorted(reason_dist.items(), key=lambda x: -x[1])
            ])
            row = self._write_table(ws, reason_df, row, 1, COLOR_HEADER)

        self._auto_col_width(ws)

    # ── 공통 헬퍼 ─────────────────────────────────────────────────────────────

    def _write_section_title(self, ws, row: int, col: int, title: str) -> None:
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, size=12)

    @staticmethod
    def _write_table(
        ws,
        df: pd.DataFrame,
        start_row: int,
        start_col: int,
        header_color: str,
    ) -> int:
        """DataFrame을 셀에 기록하고 다음 빈 행 인덱스를 반환."""
        # 헤더
        for col_offset, col_name in enumerate(df.columns):
            cell = ws.cell(
                row=start_row, column=start_col + col_offset, value=str(col_name)
            )
            PocExcelWriter._color_cell(cell, header_color)
            if header_color in (COLOR_HEADER,):
                cell.font = Font(color="FFFFFFFF", bold=True)
            else:
                cell.font = Font(bold=True)

        row = start_row + 1
        for _, data_row in df.iterrows():
            for col_offset, val in enumerate(data_row):
                # Convert numpy/pandas types to native Python for openpyxl
                if hasattr(val, "item"):
                    val = val.item()
                ws.cell(row=row, column=start_col + col_offset, value=val)
            row += 1

        return row

    @staticmethod
    def _write_kv_rows(ws, rows: list, start_row: int) -> int:
        """key-value 쌍을 2열(A, B)에 기록하고 다음 빈 행 인덱스를 반환."""
        row = start_row
        for key, val in rows:
            ws.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
            cell_val = val.item() if hasattr(val, "item") else val
            ws.cell(row=row, column=2, value=cell_val)
            row += 1
        return row

    @staticmethod
    def _color_cell(cell, hex_color: str) -> None:
        """셀 배경색 설정 (fill_type='solid')."""
        cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)

    @staticmethod
    def _auto_col_width(ws, min_width: int = 10, max_width: int = 50) -> None:
        """컬럼 너비 자동 조정."""
        for col_cells in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)
