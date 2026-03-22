"""Phase 3 테스트: Excel 작성기 (TDD)"""
import sys
from pathlib import Path
import tempfile

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))

import pytest
import pandas as pd
import numpy as np


# ─────────────────────────────────────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────────────────────────────────────

def _make_full_data():
    """PocReportData 전체 필드 채운 샘플 데이터"""
    from src.report.excel_writer import PocReportData

    by_month = pd.DataFrame({
        "month": ["1월", "2월", "3월"],
        "tp_count": [10, 12, 8],
        "fp_count": [5, 6, 4],
        "total": [15, 18, 12],
        "tp_ratio": [10/15, 12/18, 8/12],
    })

    class_imbalance = pd.DataFrame({
        "class_name": ["TP", "FP"],
        "count": [30, 15],
        "ratio": [0.667, 0.333],
    })

    fp_desc = pd.DataFrame({
        "fp_description": ["숫자나열", "타임스탬프"],
        "count": [10, 5],
        "orgs": ["CTO", "NW"],
    })

    split_comp = pd.DataFrame({
        "split_name": ["Primary", "Secondary", "Tertiary-CTO"],
        "train_n": [300, 280, 250],
        "test_n": [50, 45, 40],
        "f1_macro": [0.82, 0.79, 0.81],
        "tp_recall": [0.85, 0.80, 0.83],
        "fp_precision": [0.90, 0.88, 0.89],
        "auto_fp_coverage_at_95": [0.72, 0.68, 0.70],
        "poc_verdict": ["PASS", "PASS", "FAIL"],
    })

    curve_df = pd.DataFrame({
        "tau": [0.50, 0.60, 0.70, 0.80, 0.90, 1.00],
        "coverage": [0.95, 0.88, 0.78, 0.65, 0.50, 0.30],
        "precision": [0.85, 0.88, 0.91, 0.94, 0.96, 0.99],
        "tp_safety_rate": [0.05, 0.03, 0.02, 0.01, 0.01, 0.00],
        "auto_fp_count": [95, 88, 78, 65, 50, 30],
    })

    rule_contrib = pd.DataFrame({
        "rule_id": ["R001", "R002", "R003"],
        "hit_count": [50, 30, 10],
        "hit_rate": [0.50, 0.30, 0.10],
        "precision": [0.90, 0.85, 0.80],
        "dominant_class": ["FP-숫자나열/코드", "FP-타임스탬프", "FP-더미데이터"],
    })

    class_rule_contrib = pd.DataFrame({
        "class_name": ["FP-숫자나열/코드", "FP-타임스탬프"],
        "rule_id_count": [2, 1],
        "total_hits": [80, 30],
    })

    error_patterns = [
        ("TP", "FP", 8),
        ("FP-숫자나열/코드", "TP", 3),
        ("FP-타임스탬프", "TP", 2),
    ]

    error_samples = pd.DataFrame({
        "pk_event": [f"evt_{i:03d}" for i in range(10)],
        "actual_class": ["TP"] * 5 + ["FP"] * 5,
        "predicted_class": ["FP"] * 5 + ["TP"] * 5,
    })

    org_stats = pd.DataFrame({
        "organization": ["CTO", "NW"],
        "tp_count": [15, 10],
        "fp_count": [5, 8],
        "total": [20, 18],
        "tp_ratio": [0.75, 0.556],
    })

    class_metrics = pd.DataFrame({
        "class_name": ["TP", "FP"],
        "precision": [0.88, 0.92],
        "recall": [0.85, 0.90],
        "f1_score": [0.865, 0.909],
        "support": [30, 15],
    })

    confidence_dist = pd.DataFrame({
        "proba_range": ["0.5~0.6", "0.6~0.7", "0.7~0.8", "0.8~0.9", "0.9~1.0"],
        "count": [10, 15, 20, 30, 25],
        "ratio": [0.10, 0.15, 0.20, 0.30, 0.25],
        "cumulative_ratio": [0.10, 0.25, 0.45, 0.75, 1.00],
    })

    return PocReportData(
        data_condition="Label Only",
        split_summary={"train_n": 350, "test_n": 95, "split_method": "GroupTimeSplit"},
        poc_criteria={"passes": True, "f1_macro": 0.82, "tp_recall": 0.85, "fp_precision": 0.90},
        binary_stats={"total": {"tp": 30, "fp": 15, "total": 45, "tp_ratio": 0.667, "fp_ratio": 0.333},
                      "by_month": by_month},
        class_imbalance=class_imbalance,
        dedup_before=500,
        dedup_after=450,
        fp_description_stats=fp_desc,
        split_comparison=split_comp,
        coverage_curve={"curve": curve_df, "recommended_tau": 0.90},
        rule_contribution=rule_contrib,
        class_rule_contribution=class_rule_contrib,
        error_patterns=error_patterns,
        error_samples=error_samples,
        run_metadata={"run_datetime": "2026-03-15 10:00:00", "model_path": "models/model.joblib",
                      "data_date_range": "1월 ~ 3월"},
        business_impact={"total_fp": 15, "coverage_at_target": 0.50, "estimated_auto_fp": 7,
                         "phase1_goal_40pct_met": True},
        org_stats=org_stats,
        class_metrics=class_metrics,
        rule_vs_ml_coverage={"rule_only_coverage": 0.30, "ml_total_coverage": 0.65,
                              "ml_additional_coverage": 0.35, "overlap_count": 5, "overlap_rate": 0.25},
        error_risk_summary={"fp_to_tp_count": 2, "tp_to_fp_count": 5,
                            "total_errors": 7, "risk_rate": 0.286},
        confidence_distribution=confidence_dist,
    )


# ─────────────────────────────────────────────────────────────────────────────
# TestPocExcelWriter
# ─────────────────────────────────────────────────────────────────────────────

class TestPocExcelWriter:
    def test_ten_sheets_created(self):
        from src.report.excel_writer import PocExcelWriter

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)

        try:
            PocExcelWriter(data).write(out_path)
            from openpyxl import load_workbook
            wb = load_workbook(str(out_path))
            assert len(wb.sheetnames) == 10
        finally:
            out_path.unlink(missing_ok=True)

    def test_file_readable_by_pandas(self):
        from src.report.excel_writer import PocExcelWriter

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)

        try:
            PocExcelWriter(data).write(out_path)
            # pandas로 읽을 수 있어야 함
            xls = pd.ExcelFile(str(out_path))
            sheet_count = len(xls.sheet_names)
            xls.close()  # Windows: 파일 핸들 해제 후 삭제
            assert sheet_count == 10
        finally:
            out_path.unlink(missing_ok=True)

    def test_output_path_created(self):
        from src.report.excel_writer import PocExcelWriter

        data = _make_full_data()
        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "sub" / "poc_report.xlsx"
            PocExcelWriter(data).write(out_path)
            assert out_path.exists()


# ─────────────────────────────────────────────────────────────────────────────
# TestBuildSummarySheet
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildSummarySheet:
    def _get_sheet_values(self, data, sheet_index=0):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[sheet_index]
            values = []
            for row in ws.iter_rows(values_only=True):
                for cell_val in row:
                    if cell_val is not None:
                        values.append(str(cell_val))
            return values
        finally:
            out_path.unlink(missing_ok=True)

    def test_data_condition_row_present(self):
        data = _make_full_data()
        values = self._get_sheet_values(data, sheet_index=0)
        assert any("Label Only" in v for v in values)

    def test_poc_verdict_highlighted_green_or_red(self):
        from src.report.excel_writer import PocExcelWriter, COLOR_PASS, COLOR_FAIL
        from openpyxl import load_workbook

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[0]
            found_color = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.fgColor:
                        color_val = cell.fill.fgColor.rgb
                        # 색상 비교: 8자리 AARRGGBB 형식
                        if color_val in (COLOR_PASS, COLOR_FAIL):
                            found_color = True
            assert found_color
        finally:
            out_path.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TestBuildDataStatsSheet
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildDataStatsSheet:
    def _get_sheet_text(self, data, sheet_index=1):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[sheet_index]
            texts = []
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is not None:
                        texts.append(str(v))
            return texts
        finally:
            out_path.unlink(missing_ok=True)

    def test_tp_fp_counts_present(self):
        data = _make_full_data()
        texts = self._get_sheet_text(data)
        assert any("TP" in t for t in texts)
        assert any("FP" in t for t in texts)

    def test_monthly_distribution_row_count(self):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[1]
            # 월별 데이터 3개 월이 들어 있는지 확인
            all_values = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
            month_vals = [v for v in all_values if str(v).endswith("월")]
            assert len(month_vals) >= 3
        finally:
            out_path.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TestBuildModelPerformanceSheet
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildModelPerformanceSheet:
    def test_three_split_rows(self):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[2]
            # Primary, Secondary, Tertiary-CTO가 있는지
            all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
            assert any("Primary" in v for v in all_values)
            assert any("Secondary" in v for v in all_values)
        finally:
            out_path.unlink(missing_ok=True)

    def test_pass_fail_cell_coloring(self):
        from src.report.excel_writer import PocExcelWriter, COLOR_PASS, COLOR_FAIL
        from openpyxl import load_workbook

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[2]
            colors_found = set()
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.fgColor:
                        colors_found.add(cell.fill.fgColor.rgb)
            # PASS(녹색) 또는 FAIL(빨강) 색이 있어야 함 (8자리 AARRGGBB)
            assert colors_found & {COLOR_PASS, COLOR_FAIL}
        finally:
            out_path.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TestBuildCoverageCurveSheet
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildCoverageCurveSheet:
    def _get_all_values(self, data, sheet_index=3):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[sheet_index]
            return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        finally:
            out_path.unlink(missing_ok=True)

    def test_tau_column_present(self):
        data = _make_full_data()
        values = self._get_all_values(data)
        assert any("tau" in v.lower() for v in values)

    def test_recommended_tau_row_highlighted(self):
        from src.report.excel_writer import PocExcelWriter, COLOR_HIGHLIGHT
        from openpyxl import load_workbook

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[3]
            highlight_found = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.fgColor:
                        rgb = cell.fill.fgColor.rgb
                        if rgb == COLOR_HIGHLIGHT:  # 8자리 AARRGGBB
                            highlight_found = True
            assert highlight_found
        finally:
            out_path.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TestBuildRuleContributionSheet
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildRuleContributionSheet:
    def _get_all_values(self, data, sheet_index=4):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[sheet_index]
            return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        finally:
            out_path.unlink(missing_ok=True)

    def test_rule_id_column_present(self):
        data = _make_full_data()
        values = self._get_all_values(data)
        assert any("rule_id" in v.lower() for v in values)

    def test_sorted_by_hit_count(self):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        data = _make_full_data()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[4]
            # hit_count 컬럼(3번째)에서 숫자 값 추출
            hit_counts = []
            header_found = False
            hit_col = None
            for row in ws.iter_rows(values_only=True):
                if not header_found:
                    for ci, v in enumerate(row):
                        if v and "hit_count" in str(v).lower():
                            hit_col = ci
                            header_found = True
                    continue
                if hit_col is not None and row[hit_col] is not None:
                    try:
                        hit_counts.append(int(row[hit_col]))
                    except (ValueError, TypeError):
                        pass
            if hit_counts:
                assert hit_counts == sorted(hit_counts, reverse=True)
        finally:
            out_path.unlink(missing_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# TestBuildErrorAnalysisSheet
# ─────────────────────────────────────────────────────────────────────────────

class TestBuildErrorAnalysisSheet:
    def _get_all_values(self, data, sheet_index=5):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[sheet_index]
            return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        finally:
            out_path.unlink(missing_ok=True)

    def test_top_15_patterns_present(self):
        data = _make_full_data()
        values = self._get_all_values(data)
        # 오분류 패턴 헤더가 있어야 함
        assert any("실제" in v or "예측" in v or "건수" in v for v in values)

    def test_error_samples_present(self):
        data = _make_full_data()
        values = self._get_all_values(data)
        # 오분류 샘플의 pk_event 값이 있어야 함
        assert any("evt_" in v for v in values)


# ─────────────────────────────────────────────────────────────────────────────
# TestNewSections (Sheet 1~6 신규 섹션 + Sheet 7)
# ─────────────────────────────────────────────────────────────────────────────

class TestNewSections:
    def _write_and_load(self, data, sheet_index):
        from src.report.excel_writer import PocExcelWriter
        from openpyxl import load_workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[sheet_index]
            return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        finally:
            out_path.unlink(missing_ok=True)

    def test_summary_sheet_has_run_datetime(self):
        data = _make_full_data()
        values = self._write_and_load(data, sheet_index=0)
        assert any("2026-03-15" in v for v in values)

    def test_summary_sheet_has_estimated_auto_fp(self):
        data = _make_full_data()
        values = self._write_and_load(data, sheet_index=0)
        assert any("추정" in v or "estimated" in v.lower() or "7" in v for v in values)

    def test_data_stats_sheet_has_org_section(self):
        data = _make_full_data()
        values = self._write_and_load(data, sheet_index=1)
        # org_stats에 CTO, NW가 있어야 함
        assert any("CTO" in v for v in values)
        assert any("NW" in v for v in values)

    def test_model_perf_sheet_has_class_metrics(self):
        data = _make_full_data()
        values = self._write_and_load(data, sheet_index=2)
        assert any("class_name" in v.lower() or "precision" in v.lower() for v in values)

    def test_error_analysis_sheet_has_risk_summary(self):
        data = _make_full_data()
        values = self._write_and_load(data, sheet_index=5)
        assert any("위험" in v or "risk" in v.lower() for v in values)

    def test_confidence_sheet_exists(self):
        """Sheet 7 (index 6) 신뢰도 분포 시트가 생성되어야 한다."""
        data = _make_full_data()
        values = self._write_and_load(data, sheet_index=6)
        assert len(values) > 0
        assert any("proba" in v.lower() or "신뢰도" in v or "0.5" in v for v in values)

    def test_confidence_sheet_skip_ml_message(self):
        """confidence_distribution이 빈 DataFrame이면 메시지를 출력해야 한다."""
        from src.report.excel_writer import PocExcelWriter, PocReportData
        from openpyxl import load_workbook

        data = _make_full_data()
        data.confidence_distribution = pd.DataFrame()  # empty

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out_path = Path(f.name)
        try:
            PocExcelWriter(data).write(out_path)
            wb = load_workbook(str(out_path))
            ws = wb.worksheets[6]
            values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
            assert any("ML" in v or "skip" in v.lower() or "없음" in v for v in values)
        finally:
            out_path.unlink(missing_ok=True)
